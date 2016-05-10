from openpyxl import load_workbook
# from openpyxl.utils import coordinate_from_string, column_index_from_string
import os
# import codecs


class Database:

    def add(self, Ename, Eid, Esal, Epostype, Edate, Esuper):
        if self.searchId(self, Eid):
            return 'ID found. Try another ID'
        else:
            wb = load_workbook(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')
            ws = wb['Sheet1']

            # Find the available row to add values
            for j in range(1, 101):
                if ws.cell(row=j, column=1).value is None:
                    ws[ws.cell(row=j, column=1).coordinate] = Eid
                    ws[ws.cell(row=j, column=2).coordinate] = Ename
                    ws[ws.cell(row=j, column=3).coordinate] = Esal
                    ws[ws.cell(row=j, column=4).coordinate] = Epostype
                    ws[ws.cell(row=j, column=5).coordinate] = Edate
                    ws[ws.cell(row=j, column=6).coordinate] = Esuper
                    break

            wb.save(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')

            return 'Added worker: ' + Ename

    def edit(self, Efield, Eid, Newparameter):
        if not self.searchId(self, Eid):
            return 'ID not found'
        elif Efield == 'Name' or Efield == 'ID' or Efield == 'Salary' or Efield == 'Position' or Efield == 'Date' \
                or Efield == 'Supervisor':

            wb = load_workbook(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')
            ws = wb['Sheet1']

            fields = {'ID': 1, 'Name': 2, 'Salary': 3, 'Position': 4, 'Date': 5, 'Supervisor': 6}

             # Find the ID in the ID's column
            for i in range(1, 101):
                if ws.cell(row=i, column=1).value == Eid:
                    # Update with new value
                    ws[ws.cell(row=i, column=fields[Efield]).coordinate] = Newparameter

            wb.save(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')

            return 'Edited Field ' + Efield + ' of Worker with ID = ' + str(Eid) + ' with value: ' + str(Newparameter)
        else:
            return 'Field not valid for editing'

    def delete(self, Eid):
        if not self.searchId(self, Eid):
            return 'ID not found'
        else:
            wb = load_workbook(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')
            ws = wb['Sheet1']

            # Find the ID in first column
            for j in range(2, 101):
                if ws.cell(row=j, column=1).value == Eid:
                    for i in range(1, 7):
                        ws.cell(row=j, column=i).value = ''
                    break

            # Move last worker information to recently deleted row
            for k in range(2, 101):
                if ws.cell(row=k, column=1).value == '' and ws.cell(row=k+1, column=1).value != '':
                    for l in range(k, 101):
                        if ws.cell(row=l+1, column=1).value is None:
                            for n in range(1, 7):
                                ws.cell(row=k, column=n).value = ws.cell(row=l, column=n).value
                                ws.cell(row=l, column=n).value = None
                            break
                    break

            wb.save(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')

            return 'Worker deleted from database'

    def searchId(self, ID):
        wb = load_workbook(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')
        ws = wb['Sheet1']

        # Find the ID in the ID's column
        for j in range(2, 200):
            if ws.cell(row=j, column=1).value == str(ID):
                return True
            elif ws.cell(row=j, column=1).value == '':
                break

        return False

    def getEmployeeInfo(self, ID):
        if not self.searchId(self,ID):
            return 'ID not found'
        else:
            wb = load_workbook(os.path.dirname(os.path.dirname( __file__ )) +'/Database.xlsx')
            ws = wb['Sheet1']
            Eid = 0
            name = ''
            salary = 0
            position = ''
            date = ''
            supervisor = ''

            # Find the ID in the ID's column
            for j in range(1, 101):
                if ws.cell(row=j, column=1).value == str(ID):
                    Eid = ws.cell(row=j, column=1).value
                    name = ws.cell(row=j, column=2).value
                    salary = ws.cell(row=j, column=3).value
                    position = ws.cell(row=j, column=4).value
                    date = ws.cell(row=j, column=5).value
                    supervisor = ws.cell(row=j, column=6).value
                    break

            return {'ID': Eid, 'Name': name, 'Salary': salary, 'Position': position,
                    'Date': date, 'Supervisor': supervisor}

    def help(self):
        print('The following commands are supported by this system:\n'
              'add          Inserts a new employee to database.\n'
              'delete       Erases a specific employee from database.\n'
              'generate     Creates one of 6 different documents using the information of an employee in the database.\n'
              '             Types of documents that can be generated:\n'
              '             eCert       Employment Certification\n'
              '             eEval       Employee Performance Evaluation\n'
              '             eCont       Employment Contract\n'
              '             eDism       Employee Dismissal Letter\n'
              '             eWarn       Employee Warning Notice\n'
              '             eTrain      Employee Training Certificate\n'
              'print        Prints selected document.\n'
              'email        Sends an email with a specified document attached to an given email address.\n'
              '\n'
              'Examples of command usage:\n'
              'Formats:  \n'
              'add Name ID Salary Position Date Supervisor\n'
              '\n'
              'delete ID\n'
              '\n'
              'generate eCert ID Output_File_Name\n'
              'generate eEval ID Output_File_Name\n'
              'generate eCont ID Output_File_Name\n'
              'generate eDism ID Output_File_Name\n'
              'generate eWarn ID Output_File_Name\n'
              'generate eTrain ID Output_File_Name\n'
              '\n'
              'print fileName\n'
              '\n'
              'email email_Address fileName\n')


